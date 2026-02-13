# app/core/excel_utils.py

from __future__ import annotations

from dataclasses import dataclass
from typing import Any, Iterable, Optional

from fastapi import HTTPException, UploadFile
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class ExtractEmailsResult:
    emails: list[str]                 # valid + normalized + unique (within file)
    total_rows: int                   # rows processed (excluding header)
    extracted_values: int             # non-empty values read from email column
    invalid_count: int
    sample_invalid: list[str]
    headers: list[str]                # normalized headers found in sheet
    used_email_header: str            # which header matched
    used_sheet_name: str              # sheet name used


def norm_header(x: Any) -> str:
    if x is None:
        return ""
    return (
        str(x)
        .strip()
        .lower()
        .replace(" ", "_")
        .replace("-", "_")
    )


def norm_email(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip().lower()


def looks_like_email(s: str) -> bool:
    # validation خفيفة وسريعة (مش RFC كامل)
    if not s or "@" not in s:
        return False
    local, _, domain = s.partition("@")
    if not local or not domain or "." not in domain:
        return False
    if " " in s:
        return False
    return True


def _select_worksheet(wb, sheet_name: Optional[str]) -> Worksheet:
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise HTTPException(status_code=422, detail=f"Sheet '{sheet_name}' not found in the Excel file")
        ws: Worksheet = wb[sheet_name]
    else:
        ws: Worksheet = wb.active

    if ws is None:
        raise HTTPException(status_code=422, detail="No worksheet found in the Excel file")
    return ws


def detect_email_column_index(headers: list[str], common_headers: Iterable[str], hint: Optional[str]) -> tuple[int, str]:
    """
    Returns (index, matched_header_name).
    - hint: a header name from frontend (optional).
    - common_headers: list of acceptable header variants.
    """
    hint_norm = norm_header(hint) if hint else ""

    if hint_norm and hint_norm in headers:
        return headers.index(hint_norm), hint_norm

    common_set = {norm_header(h) for h in common_headers}
    for i, h in enumerate(headers):
        if h in common_set:
            return i, h

    found = [h for h in headers if h]
    raise HTTPException(
        status_code=422,
        detail=f"Could not find an email column. Found columns: {found}. "
               f"Expected something like: {sorted(common_set)}",
    )


def extract_emails_from_xlsx(
    *,
    file: UploadFile,
    sheet_name: Optional[str],
    email_column_hint: Optional[str],
    common_email_headers: Iterable[str],
    max_samples: int = 20,
) -> ExtractEmailsResult:
    # Validate file extension (خفيف)
    filename = (file.filename or "").lower()
    if not filename.endswith(".xlsx"):
        raise HTTPException(status_code=422, detail="Invalid file type. Please upload an .xlsx file")

    # Read Excel
    try:
        wb = load_workbook(file.file, read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=422, detail="Failed to read Excel file. Make sure it is a valid .xlsx") from e

    ws = _select_worksheet(wb, sheet_name)
    used_sheet_name = ws.title

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        raise HTTPException(status_code=422, detail="Excel file is empty")

    headers_raw = list(header_row) if header_row else []
    headers = [norm_header(h) for h in headers_raw]

    email_col_idx, matched_header = detect_email_column_index(
        headers=headers,
        common_headers=common_email_headers,
        hint=email_column_hint,
    )

    # Extract + validate + dedupe داخل الملف
    total_rows = 0
    extracted_values = 0
    invalid_count = 0
    sample_invalid: list[str] = []

    seen: set[str] = set()
    emails: list[str] = []

    for row in rows_iter:
        total_rows += 1
        cell_val = row[email_col_idx] if row and len(row) > email_col_idx else None
        if cell_val is None:
            continue

        extracted_values += 1
        email = norm_email(cell_val)

        if not looks_like_email(email):
            invalid_count += 1
            if len(sample_invalid) < max_samples:
                sample_invalid.append(email)
            continue

        if email in seen:
            continue

        seen.add(email)
        emails.append(email)

    if not emails and extracted_values == 0:
        raise HTTPException(status_code=422, detail="No emails found in the uploaded file")

    return ExtractEmailsResult(
        emails=emails,
        total_rows=total_rows,
        extracted_values=extracted_values,
        invalid_count=invalid_count,
        sample_invalid=sample_invalid,
        headers=[h for h in headers if h],
        used_email_header=matched_header,
        used_sheet_name=used_sheet_name,
    )
