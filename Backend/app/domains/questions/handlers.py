from __future__ import annotations

from typing import Any, Mapping

import uuid

from sqlalchemy import text
from sqlalchemy.orm import Session

from io import BytesIO
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage

from app.core.background_jobs.registry import register_handler
from app.core.supabase_client import supabase
from app.core.config import settings


def validate_and_normalize_question_payload(payload) -> dict[str, Any]:
    question_type = (payload.type or "").strip().lower()

    if not question_type:
        return {"ok": False, "status_code": 422, "detail": "type is required"}

    handlers = {
        "multiple_choice": validate_and_normalize_multiple_choice,
        "multi_select": validate_and_normalize_multi_select,
        "true_false": validate_and_normalize_true_false,
        "short_answer": validate_and_normalize_short_answer,
        "essay": validate_and_normalize_essay,
        "code": validate_and_normalize_code
    }

    handler = handlers.get(question_type)
    if not handler:
        return {
            "ok": False,
            "status_code": 422,
            "detail": f"Unsupported question type: {question_type}"
        }

    return handler(payload)


def validate_and_normalize_multiple_choice(payload) -> dict[str, Any]:
    question_text = (payload.question_text or "").strip()
    if not question_text:
        return {"ok": False, "status_code": 422, "detail": "question_text is required"}

    difficulty = (payload.difficulty or "").strip().lower()
    if not difficulty:
        return {"ok": False, "status_code": 422, "detail": "difficulty is required"}

    options = payload.options
    if not isinstance(options, list) or len(options) < 2:
        return {
            "ok": False,
            "status_code": 422,
            "detail": "multiple_choice questions must include at least 2 options"
        }

    normalized_options = []
    seen_ids = set()

    for option in options:
        option_id = (option.id or "").strip()
        option_text = (option.text or "").strip()

        if not option_id:
            return {
                "ok": False,
                "status_code": 422,
                "detail": "Each option must include a non-empty id"
            }

        if not option_text:
            return {
                "ok": False,
                "status_code": 422,
                "detail": "Each option must include non-empty text"
            }

        if option_id in seen_ids:
            return {
                "ok": False,
                "status_code": 422,
                "detail": "Option ids must be unique"
            }

        seen_ids.add(option_id)
        normalized_options.append({
            "id": option_id,
            "text": option_text,
        })

    expected_answer = (payload.expected_answer or "").strip()
    if not expected_answer:
        return {
            "ok": False,
            "status_code": 422,
            "detail": "expected_answer is required for multiple_choice questions"
        }

    if expected_answer not in seen_ids:
        return {
            "ok": False,
            "status_code": 422,
            "detail": "expected_answer must match one of the option ids"
        }

    explanation = payload.explanation
    if isinstance(explanation, str):
        explanation = explanation.strip() or None

    return {
        "ok": True,
        "data": {
            "question_text": question_text,
            "type": question_type_from_payload(payload),
            "difficulty": difficulty,
            "explanation": explanation,
            "options": normalized_options,
            "expected_answer": expected_answer,
            "grading_rubric": None,
            "max_score": 1,
            "auto_gradable": True,
            "source": "manual",
            "approval_status": "approved",
            "usage_count": 0,
            "success_rate": None,
            "average_time_seconds": None,
            "tags": None,
        }
    }



def validate_and_normalize_multi_select(payload) -> dict[str, Any]:
    question_text = (payload.question_text or "").strip()
    if not question_text:
        return {"ok": False, "status_code": 422, "detail": "question_text is required"}

    difficulty = (payload.difficulty or "").strip().lower()
    if not difficulty:
        return {"ok": False, "status_code": 422, "detail": "difficulty is required"}

    options = payload.options
    if not isinstance(options, list) or len(options) < 2:
        return {
            "ok": False,
            "status_code": 422,
            "detail": "multi_select questions must include at least 2 options"
        }

    normalized_options = []
    seen_ids = set()

    for option in options:
        option_id = (option.id or "").strip()
        option_text = (option.text or "").strip()

        if not option_id:
            return {
                "ok": False,
                "status_code": 422,
                "detail": "Each option must include a non-empty id"
            }

        if not option_text:
            return {
                "ok": False,
                "status_code": 422,
                "detail": "Each option must include non-empty text"
            }

        if option_id in seen_ids:
            return {
                "ok": False,
                "status_code": 422,
                "detail": "Option ids must be unique"
            }

        seen_ids.add(option_id)
        normalized_options.append({
            "id": option_id,
            "text": option_text,
        })

    expected_answer = payload.expected_answer
    if not isinstance(expected_answer, list) or not expected_answer:
        return {
            "ok": False,
            "status_code": 422,
            "detail": "expected_answer is required for multi_select questions and must be a non-empty list"
        }

    normalized_expected_answer = []
    seen_answer_ids = set()

    for answer_id in expected_answer:
        if not isinstance(answer_id, str):
            return {
                "ok": False,
                "status_code": 422,
                "detail": "Each expected_answer item must be a string"
            }

        answer_id = answer_id.strip()
        if not answer_id:
            return {
                "ok": False,
                "status_code": 422,
                "detail": "Each expected_answer item must be non-empty"
            }

        if answer_id not in seen_ids:
            return {
                "ok": False,
                "status_code": 422,
                "detail": "Each expected_answer item must match one of the option ids"
            }

        if answer_id in seen_answer_ids:
            return {
                "ok": False,
                "status_code": 422,
                "detail": "expected_answer must not contain duplicates"
            }

        seen_answer_ids.add(answer_id)
        normalized_expected_answer.append(answer_id)

    explanation = payload.explanation
    if isinstance(explanation, str):
        explanation = explanation.strip() or None

    return {
        "ok": True,
        "data": {
            "question_text": question_text,
            "type": question_type_from_payload(payload),
            "difficulty": difficulty,
            "explanation": explanation,
            "options": normalized_options,
            "expected_answer": normalized_expected_answer,
            "grading_rubric": None,
            "max_score": 1,
            "auto_gradable": True,
            "source": "manual",
            "approval_status": "approved",
            "usage_count": 0,
            "success_rate": None,
            "average_time_seconds": None,
            "tags": None,
        }
    }



def validate_and_normalize_true_false(payload) -> dict[str, Any]:
    question_text = (payload.question_text or "").strip()
    if not question_text:
        return {"ok": False, "status_code": 422, "detail": "question_text is required"}

    difficulty = (payload.difficulty or "").strip().lower()
    if not difficulty:
        return {"ok": False, "status_code": 422, "detail": "difficulty is required"}

    expected_answer = payload.expected_answer
    if not isinstance(expected_answer, str):
        return {
            "ok": False,
            "status_code": 422,
            "detail": "expected_answer is required for true_false questions and must be a string"
        }

    expected_answer = expected_answer.strip().lower()
    if expected_answer not in {"true", "false"}:
        return {
            "ok": False,
            "status_code": 422,
            "detail": "expected_answer for true_false must be either 'true' or 'false'"
        }

    explanation = payload.explanation
    if isinstance(explanation, str):
        explanation = explanation.strip() or None

    normalized_options = [
        {"id": "true", "text": "True"},
        {"id": "false", "text": "False"},
    ]

    return {
        "ok": True,
        "data": {
            "question_text": question_text,
            "type": question_type_from_payload(payload),
            "difficulty": difficulty,
            "explanation": explanation,
            "options": normalized_options,
            "expected_answer": expected_answer,
            "grading_rubric": None,
            "max_score": 1,
            "auto_gradable": True,
            "source": "manual",
            "approval_status": "approved",
            "usage_count": 0,
            "success_rate": None,
            "average_time_seconds": None,
            "tags": None,
        }
    }



def validate_and_normalize_short_answer(payload) -> dict[str, Any]:
    question_text = (payload.question_text or "").strip()
    if not question_text:
        return {"ok": False, "status_code": 422, "detail": "question_text is required"}

    difficulty = (payload.difficulty or "").strip().lower()
    if not difficulty:
        return {"ok": False, "status_code": 422, "detail": "difficulty is required"}

    expected_answer = payload.expected_answer
    if not isinstance(expected_answer, str):
        return {
            "ok": False,
            "status_code": 422,
            "detail": "expected_answer is required for short_answer questions and must be a string"
        }

    expected_answer = expected_answer.strip()
    if not expected_answer:
        return {
            "ok": False,
            "status_code": 422,
            "detail": "expected_answer is required for short_answer questions"
        }

    explanation = payload.explanation
    if isinstance(explanation, str):
        explanation = explanation.strip() or None

    grading_rubric = getattr(payload, "grading_rubric", None)
    if grading_rubric is not None and not isinstance(grading_rubric, dict):
        return {
            "ok": False,
            "status_code": 422,
            "detail": "grading_rubric must be an object when provided"
        }

    return {
        "ok": True,
        "data": {
            "question_text": question_text,
            "type": question_type_from_payload(payload),
            "difficulty": difficulty,
            "explanation": explanation,
            "options": None,
            "expected_answer": expected_answer,
            "grading_rubric": grading_rubric,
            "max_score": 1,
            "auto_gradable": False,
            "source": "manual",
            "approval_status": "approved",
            "usage_count": 0,
            "success_rate": None,
            "average_time_seconds": None,
            "tags": None,
        }
    }



def validate_and_normalize_essay(payload) -> dict[str, Any]:
    question_text = (payload.question_text or "").strip()
    if not question_text:
        return {"ok": False, "status_code": 422, "detail": "question_text is required"}

    difficulty = (payload.difficulty or "").strip().lower()
    if not difficulty:
        return {"ok": False, "status_code": 422, "detail": "difficulty is required"}

    expected_answer = payload.expected_answer
    if expected_answer is not None:
        if not isinstance(expected_answer, str):
            return {
                "ok": False,
                "status_code": 422,
                "detail": "expected_answer must be a string when provided for essay questions"
            }

        expected_answer = expected_answer.strip() or None

    explanation = payload.explanation
    if isinstance(explanation, str):
        explanation = explanation.strip() or None

    grading_rubric = getattr(payload, "grading_rubric", None)
    if grading_rubric is not None and not isinstance(grading_rubric, dict):
        return {
            "ok": False,
            "status_code": 422,
            "detail": "grading_rubric must be an object when provided"
        }

    return {
        "ok": True,
        "data": {
            "question_text": question_text,
            "type": question_type_from_payload(payload),
            "difficulty": difficulty,
            "explanation": explanation,
            "options": None,
            "expected_answer": expected_answer,
            "grading_rubric": grading_rubric,
            "max_score": 1,
            "auto_gradable": False,
            "source": "manual",
            "approval_status": "approved",
            "usage_count": 0,
            "success_rate": None,
            "average_time_seconds": None,
            "tags": None,
        }
    }



def validate_and_normalize_code(payload) -> dict[str, Any]:
    question_text = (payload.question_text or "").strip()
    if not question_text:
        return {"ok": False, "status_code": 422, "detail": "question_text is required"}

    difficulty = (payload.difficulty or "").strip().lower()
    if not difficulty:
        return {"ok": False, "status_code": 422, "detail": "difficulty is required"}

    expected_answer = payload.expected_answer
    if not isinstance(expected_answer, dict):
        return {
            "ok": False,
            "status_code": 422,
            "detail": "expected_answer is required for code questions and must be an object"
        }

    code = expected_answer.get("code")
    if not isinstance(code, str) or not code.strip():
        return {
            "ok": False,
            "status_code": 422,
            "detail": "expected_answer.code is required and must be a non-empty string"
        }

    language = expected_answer.get("language")
    if language is not None:
        if not isinstance(language, str) or not language.strip():
            return {
                "ok": False,
                "status_code": 422,
                "detail": "expected_answer.language must be a non-empty string when provided"
            }
        language = language.strip().lower()

    normalized_expected_answer = {
        "code": code.strip(),
        "language": language,
    }

    explanation = payload.explanation
    if isinstance(explanation, str):
        explanation = explanation.strip() or None

    grading_rubric = getattr(payload, "grading_rubric", None)
    if grading_rubric is not None and not isinstance(grading_rubric, dict):
        return {
            "ok": False,
            "status_code": 422,
            "detail": "grading_rubric must be an object when provided"
        }

    return {
        "ok": True,
        "data": {
            "question_text": question_text,
            "type": question_type_from_payload(payload),
            "difficulty": difficulty,
            "explanation": explanation,
            "options": None,
            "expected_answer": normalized_expected_answer,
            "grading_rubric": grading_rubric,
            "max_score": 1,
            "auto_gradable": False,
            "source": "manual",
            "approval_status": "approved",
            "usage_count": 0,
            "success_rate": None,
            "average_time_seconds": None,
            "tags": None,
        }
    }



def question_type_from_payload(payload) -> str:
    return (payload.type or "").strip().lower()




QUESTION_BANK_EXPORT_TEMPLATE_PATH = "assets/question_bank_export_template.xlsx"
CODE_FONT = Font(name="Consolas", size=10)

# =========================
# Metadata helper — called once per sheet by the orchestrator
# =========================
def _fill_course_metadata(*, ws, course_row: Mapping, total_questions: int) -> None:
    ws["D2"] = course_row["title"]
    ws["G2"] = datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")
    ws["D4"] = course_row["course_code"] or ""
    ws["G4"] = total_questions


# =========================
# Multiple Choice sheet handler
# =========================
def _handle_multiple_choice_questions(*, db: Session, ws, course_id: int, bucket: str) -> None:
    rows = db.execute(
        text("""
            SELECT q.id, q.difficulty, q.question_text, q.explanation,
                   q.options, q.expected_answer, q.image_key,
                   t.title AS topic_title, t.description AS topic_description
            FROM questions q
            JOIN topics t ON t.id = q.topic_id
            WHERE q.course_id = :course_id
              AND q.approval_status = 'approved'
              AND q.type = 'multiple_choice'
            ORDER BY t.id, q.id
        """),
        {"course_id": course_id},
    ).mappings().all()

    current_row = 8
    for q in rows:
        options = q["options"] or []
        text_by_id = {opt.get("id"): opt.get("text", "") for opt in options}

        ws.cell(row=current_row, column=1, value=q["topic_title"])
        ws.cell(row=current_row, column=2, value=q["topic_description"])
        ws.cell(row=current_row, column=3, value=q["id"])
        ws.cell(row=current_row, column=4, value="multiple_choice")
        ws.cell(row=current_row, column=5, value=q["difficulty"])
        ws.cell(row=current_row, column=6, value=q["question_text"])

        option_cells = [opt.get("text", "") for opt in options[:4]]
        while len(option_cells) < 4:
            option_cells.append("")
        for i, val in enumerate(option_cells):
            ws.cell(row=current_row, column=7 + i, value=val)

        ws.cell(row=current_row, column=11, value=text_by_id.get(q["expected_answer"], str(q["expected_answer"])))
        ws.cell(row=current_row, column=12, value=q["explanation"] or "")

        if q["image_key"]:
            try:
                image_bytes = supabase.storage.from_(bucket).download(q["image_key"])
                img = XLImage(BytesIO(image_bytes))
                img.width = 120
                img.height = 120
                ws.add_image(img, f"M{current_row}")
                ws.row_dimensions[current_row].height = 90
            except Exception:
                pass

        current_row += 1


# =========================
# Multi Select sheet handler
# =========================
def _handle_multi_select_questions(*, db: Session, ws, course_id: int, bucket: str) -> None:
    rows = db.execute(
        text("""
            SELECT q.id, q.difficulty, q.question_text, q.explanation,
                   q.options, q.expected_answer, q.image_key,
                   t.title AS topic_title, t.description AS topic_description
            FROM questions q
            JOIN topics t ON t.id = q.topic_id
            WHERE q.course_id = :course_id
              AND q.approval_status = 'approved'
              AND q.type = 'multi_select'
            ORDER BY t.id, q.id
        """),
        {"course_id": course_id},
    ).mappings().all()

    current_row = 8
    for q in rows:
        options = q["options"] or []
        text_by_id = {opt.get("id"): opt.get("text", "") for opt in options}

        ws.cell(row=current_row, column=1, value=q["topic_title"])
        ws.cell(row=current_row, column=2, value=q["topic_description"])
        ws.cell(row=current_row, column=3, value=q["id"])
        ws.cell(row=current_row, column=4, value="multi_select")
        ws.cell(row=current_row, column=5, value=q["difficulty"])
        ws.cell(row=current_row, column=6, value=q["question_text"])

        option_cells = [opt.get("text", "") for opt in options[:4]]
        while len(option_cells) < 4:
            option_cells.append("")
        for i, val in enumerate(option_cells):
            ws.cell(row=current_row, column=7 + i, value=val)

        expected = q["expected_answer"] or []
        ws.cell(row=current_row, column=11, value=", ".join(text_by_id.get(i, str(i)) for i in expected))
        ws.cell(row=current_row, column=12, value=q["explanation"] or "")

        if q["image_key"]:
            try:
                image_bytes = supabase.storage.from_(bucket).download(q["image_key"])
                img = XLImage(BytesIO(image_bytes))
                img.width = 120
                img.height = 120
                ws.add_image(img, f"M{current_row}")
                ws.row_dimensions[current_row].height = 90
            except Exception:
                pass

        current_row += 1


# =========================
# True/False sheet handler (no option columns)
# =========================
def _handle_true_false_questions(*, db: Session, ws, course_id: int, bucket: str) -> None:
    rows = db.execute(
        text("""
            SELECT q.id, q.difficulty, q.question_text, q.explanation,
                   q.expected_answer, q.image_key,
                   t.title AS topic_title, t.description AS topic_description
            FROM questions q
            JOIN topics t ON t.id = q.topic_id
            WHERE q.course_id = :course_id
              AND q.approval_status = 'approved'
              AND q.type = 'true_false'
            ORDER BY t.id, q.id
        """),
        {"course_id": course_id},
    ).mappings().all()

    current_row = 8
    for q in rows:
        ws.cell(row=current_row, column=1, value=q["topic_title"])
        ws.cell(row=current_row, column=2, value=q["topic_description"])
        ws.cell(row=current_row, column=3, value=q["id"])
        ws.cell(row=current_row, column=4, value="true_false")
        ws.cell(row=current_row, column=5, value=q["difficulty"])
        ws.cell(row=current_row, column=6, value=q["question_text"])

        ws.cell(row=current_row, column=7, value="True" if str(q["expected_answer"]).lower() == "true" else "False")
        ws.cell(row=current_row, column=8, value=q["explanation"] or "")

        if q["image_key"]:
            try:
                image_bytes = supabase.storage.from_(bucket).download(q["image_key"])
                img = XLImage(BytesIO(image_bytes))
                img.width = 120
                img.height = 120
                ws.add_image(img, f"I{current_row}")
                ws.row_dimensions[current_row].height = 90
            except Exception:
                pass

        current_row += 1


# =========================
# Essay sheet handler (no option columns; grading_rubric is not exported)
# =========================
def _handle_essay_questions(*, db: Session, ws, course_id: int, bucket: str) -> None:
    rows = db.execute(
        text("""
            SELECT q.id, q.difficulty, q.question_text, q.explanation,
                   q.expected_answer, q.image_key,
                   t.title AS topic_title, t.description AS topic_description
            FROM questions q
            JOIN topics t ON t.id = q.topic_id
            WHERE q.course_id = :course_id
              AND q.approval_status = 'approved'
              AND q.type = 'essay'
            ORDER BY t.id, q.id
        """),
        {"course_id": course_id},
    ).mappings().all()

    current_row = 8
    for q in rows:
        ws.cell(row=current_row, column=1, value=q["topic_title"])
        ws.cell(row=current_row, column=2, value=q["topic_description"])
        ws.cell(row=current_row, column=3, value=q["id"])
        ws.cell(row=current_row, column=4, value="essay")
        ws.cell(row=current_row, column=5, value=q["difficulty"])
        ws.cell(row=current_row, column=6, value=q["question_text"])

        ws.cell(row=current_row, column=7, value=q["expected_answer"] or "")
        ws.cell(row=current_row, column=8, value=q["explanation"] or "")

        if q["image_key"]:
            try:
                image_bytes = supabase.storage.from_(bucket).download(q["image_key"])
                img = XLImage(BytesIO(image_bytes))
                img.width = 120
                img.height = 120
                ws.add_image(img, f"I{current_row}")
                ws.row_dimensions[current_row].height = 90
            except Exception:
                pass

        current_row += 1


# =========================
# Short Answer sheet handler (no option columns)
# =========================
def _handle_short_answer_questions(*, db: Session, ws, course_id: int, bucket: str) -> None:
    rows = db.execute(
        text("""
            SELECT q.id, q.difficulty, q.question_text, q.explanation,
                   q.expected_answer, q.image_key,
                   t.title AS topic_title, t.description AS topic_description
            FROM questions q
            JOIN topics t ON t.id = q.topic_id
            WHERE q.course_id = :course_id
              AND q.approval_status = 'approved'
              AND q.type = 'short_answer'
            ORDER BY t.id, q.id
        """),
        {"course_id": course_id},
    ).mappings().all()

    current_row = 8
    for q in rows:
        ws.cell(row=current_row, column=1, value=q["topic_title"])
        ws.cell(row=current_row, column=2, value=q["topic_description"])
        ws.cell(row=current_row, column=3, value=q["id"])
        ws.cell(row=current_row, column=4, value="short_answer")
        ws.cell(row=current_row, column=5, value=q["difficulty"])
        ws.cell(row=current_row, column=6, value=q["question_text"])

        ws.cell(row=current_row, column=7, value=q["expected_answer"] or "")
        ws.cell(row=current_row, column=8, value=q["explanation"] or "")

        if q["image_key"]:
            try:
                image_bytes = supabase.storage.from_(bucket).download(q["image_key"])
                img = XLImage(BytesIO(image_bytes))
                img.width = 120
                img.height = 120
                ws.add_image(img, f"I{current_row}")
                ws.row_dimensions[current_row].height = 90
            except Exception:
                pass

        current_row += 1


# =========================
# Code sheet handler (no option columns; monospace font on code content)
# =========================
def _handle_code_questions(*, db: Session, ws, course_id: int, bucket: str) -> None:
    rows = db.execute(
        text("""
            SELECT q.id, q.difficulty, q.question_text, q.explanation,
                   q.expected_answer, q.image_key,
                   t.title AS topic_title, t.description AS topic_description
            FROM questions q
            JOIN topics t ON t.id = q.topic_id
            WHERE q.course_id = :course_id
              AND q.approval_status = 'approved'
              AND q.type = 'code'
            ORDER BY t.id, q.id
        """),
        {"course_id": course_id},
    ).mappings().all()

    current_row = 8
    for q in rows:
        ws.cell(row=current_row, column=1, value=q["topic_title"])
        ws.cell(row=current_row, column=2, value=q["topic_description"])
        ws.cell(row=current_row, column=3, value=q["id"])
        ws.cell(row=current_row, column=4, value="code")
        ws.cell(row=current_row, column=5, value=q["difficulty"])

        question_text_cell = ws.cell(row=current_row, column=6, value=q["question_text"])

        expected = q["expected_answer"]
        if isinstance(expected, dict):
            language = expected.get("language")
            code = expected.get("code", "")
            expected_value = f"# {language}\n{code}" if language else code
        else:
            expected_value = str(expected) if expected is not None else ""

        expected_answer_cell = ws.cell(row=current_row, column=7, value=expected_value)
        ws.cell(row=current_row, column=8, value=q["explanation"] or "")

        question_text_cell.font = CODE_FONT
        expected_answer_cell.font = CODE_FONT

        if q["image_key"]:
            try:
                image_bytes = supabase.storage.from_(bucket).download(q["image_key"])
                img = XLImage(BytesIO(image_bytes))
                img.width = 120
                img.height = 120
                ws.add_image(img, f"I{current_row}")
                ws.row_dimensions[current_row].height = 90
            except Exception:
                pass

        current_row += 1


# =========================
# Orchestrator — registered as the "question_bank_export" job handler
# =========================
def question_bank_xlsx_export_handler(*, db: Session, payload: dict) -> dict:
    course_id = payload["course_id"]

    # =========================
    # 1) Validate course still exists
    # =========================
    course_row = db.execute(
        text("""
            SELECT id, title, course_code
            FROM courses
            WHERE id = :course_id
            LIMIT 1
        """),
        {"course_id": course_id},
    ).mappings().first()

    if not course_row:
        raise ValueError(f"Course {course_id} not found")

    # =========================
    # 2) Total approved question count (shown in every sheet's metadata)
    # =========================
    total_row = db.execute(
        text("""
            SELECT COUNT(*) AS total
            FROM questions
            WHERE course_id = :course_id
              AND approval_status = 'approved'
        """),
        {"course_id": course_id},
    ).mappings().first()
    total_questions = total_row["total"] if total_row else 0

    # =========================
    # 3) Load template, fill metadata on every sheet
    # =========================
    wb = load_workbook(QUESTION_BANK_EXPORT_TEMPLATE_PATH)
    bucket = settings.supabase_private_bucket

    for sheet_name in wb.sheetnames:
        _fill_course_metadata(ws=wb[sheet_name], course_row=course_row, total_questions=total_questions)

    # =========================
    # 4) Dispatch each question type to its own sheet handler
    # =========================
    _handle_multiple_choice_questions(db=db, ws=wb["Multiple Choice"], course_id=course_id, bucket=bucket)
    _handle_multi_select_questions(db=db, ws=wb["Multi Select"], course_id=course_id, bucket=bucket)
    _handle_true_false_questions(db=db, ws=wb["True False"], course_id=course_id, bucket=bucket)
    _handle_essay_questions(db=db, ws=wb["Essay"], course_id=course_id, bucket=bucket)
    _handle_short_answer_questions(db=db, ws=wb["Short Answer"], course_id=course_id, bucket=bucket)
    _handle_code_questions(db=db, ws=wb["Code"], course_id=course_id, bucket=bucket)

    # =========================
    # 5) Save in-memory, upload to Supabase
    # =========================
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    storage_key = f"courses/{course_id}/exports/{uuid.uuid4()}.xlsx"
    supabase.storage.from_(bucket).upload(
        storage_key,
        buffer.getvalue(),
        {"content-type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    )

    return {"storage_key": storage_key, "file_size_bytes": buffer.getbuffer().nbytes}
